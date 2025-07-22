"""Simplified Excel report generation for ND2 analysis results."""

import pandas as pd
import os
import logging
from typing import Dict, List

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available. Install with: pip install openpyxl")

from config import EXCEL_SETTINGS
from data_models import ProcessingResults

logger = logging.getLogger(__name__)

class ExcelReporter:
    """Simple Excel report generator for ND2 analysis results."""
    
    def __init__(self):
        """Initialize Excel reporter with simple styling."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel functionality. Activate your venv and install via pip install openpyxl.")
        
        # Simple styling
        self.header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        # Highlight fill for representative image rows
        self.highlight_fill = PatternFill(
            start_color=EXCEL_SETTINGS['highlight_color'],
            end_color=EXCEL_SETTINGS['highlight_color'],
            fill_type='solid'
        )
    
    def create_simple_report(self, results: ProcessingResults, output_path: str) -> bool:
        """
        Create a simple 3-sheet Excel report.
        
        Args:
            results: Processing results containing all data
            output_path: Path to save Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create workbook with 3 simple sheets
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Sheet 1: Mouse Averages (main results)
            self._create_mouse_averages_sheet(wb, results.group_summary)
            
            # Sheet 2: Raw Data with highlighted representative images
            self._create_raw_data_sheet(wb, results.raw_data, results.representative_images)
            
            # Sheet 3: Summary Info
            self._create_summary_sheet(wb, results)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Simple Excel report created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating Excel report: {e}")
            return False
    
    def _create_mouse_averages_sheet(self, wb: openpyxl.Workbook, mouse_data: pd.DataFrame) -> None:
        """Create mouse averages sheet - the main results."""
        ws = wb.create_sheet("Mouse_Averages")
        
        # Add headers
        headers = list(mouse_data.columns)
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
        
        # Add data
        for _, row in mouse_data.iterrows():
            ws.append(row.tolist())
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        logger.info("Mouse averages sheet created")
    
    def _create_raw_data_sheet(
        self,
        wb: openpyxl.Workbook,
        raw_data: pd.DataFrame,
        rep_images: Dict[str, List[str]]
    ) -> None:
        """Create raw data sheet with all individual measurements."""
        ws = wb.create_sheet("Raw_Data")
        
        # Add headers
        headers = list(raw_data.columns)
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
        
        # Add data rows
        for _, row in raw_data.iterrows():
            ws.append(row.tolist())
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        # Highlight rows matching representative image filenames
        rep_files = set(sum(rep_images.values(), []))
        headers = [c.value for c in ws[1]]
        if 'Filename' in headers:
            idx = headers.index('Filename') + 1  # Excel columns are 1-based
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[idx-1]
                if cell.value in rep_files:
                    for c in row:
                        c.fill = self.highlight_fill
        
        logger.info("Raw data sheet created")
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, results: ProcessingResults) -> None:
        """Create summary information sheet."""
        ws = wb.create_sheet("Summary")
        
        # Analysis summary
        ws.append(["Analysis Summary"])
        ws.append([])
        
        # Basic statistics
        total_images = len(results.raw_data)
        total_mice = len(results.raw_data['MouseID'].unique())
        total_groups = len(results.raw_data['Group'].unique())
        
        ws.append(["Total Images Analyzed", total_images])
        ws.append(["Total Mice", total_mice])
        ws.append(["Total Groups", total_groups])
        ws.append([])
        
        # Groups breakdown
        ws.append(["Group Breakdown"])
        group_counts = results.raw_data.groupby(['Group', 'MouseID']).size().groupby('Group').size()
        for group, count in group_counts.items():
            ws.append([group, f"{count} mice"])
        
        # Style the summary title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A6'].font = Font(bold=True)
        
        logger.info("Summary sheet created")
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths to fit content."""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            if column_letter:
                # Set width with reasonable limits
                adjusted_width = min(max(max_length + 2, 10), 40)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_to_csv(self, results: ProcessingResults, output_dir: str) -> Dict[str, str]:
        """
        Export results to CSV files as backup/alternative.
        
        Args:
            results: Processing results
            output_dir: Directory to save CSV files
            
        Returns:
            Dictionary of CSV file paths
        """
        os.makedirs(output_dir, exist_ok=True)
        
        csv_files = {}
        
        # Mouse averages (main results)
        mouse_path = os.path.join(output_dir, "mouse_averages.csv")
        results.group_summary.to_csv(mouse_path, index=False)
        csv_files['mouse_averages'] = mouse_path
        
        # Raw data
        raw_path = os.path.join(output_dir, "raw_data.csv")
        results.raw_data.to_csv(raw_path, index=False)
        csv_files['raw_data'] = raw_path
        
        logger.info(f"Exported {len(csv_files)} CSV files")
        return csv_files

# Backwards compatibility function
def create_excel_report(metrics_list, output_path, representative_images=None):
    """Legacy function for backwards compatibility."""
    logger.warning("create_excel_report is deprecated. Use ExcelReporter.create_simple_report instead.")
    
    # Convert old format to new format
    if metrics_list:
        raw_data = pd.DataFrame([m.to_dict() for m in metrics_list])
        from image_processing import calculate_group_statistics
        group_summary = calculate_group_statistics(raw_data)
        
        # Create minimal results object
        class SimpleResults:
            def __init__(self, raw_data, group_summary):
                self.raw_data = raw_data
                self.group_summary = group_summary
        
        results = SimpleResults(raw_data, group_summary)
        reporter = ExcelReporter()
        return reporter.create_simple_report(results, output_path)
    
    return False
